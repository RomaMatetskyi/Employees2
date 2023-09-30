import csv  # Для роботи з CSV-файлами
import openpyxl  # Для роботи з Excel-файлами
from datetime import datetime, date  # Для роботи з датами

# Функція для розрахунку віку на основі дати народження
def calculate_age(birthdate):
    today = date.today()
    age = today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))
    return age

# Створюємо робочу книгу XLSX
wb = openpyxl.Workbook()
try:
    ws_all = wb.active
    ws_all.title = "all"  # Задаємо назву активного аркуша
except Exception as e:
    print(f"Помилка створення аркушу 'all': {e}")
    exit(1)

# Додаємо заголовки до аркушу "all"
headers = ['Прізвище', 'Ім’я', 'По батькові', 'Дата народження', 'Вік']
ws_all.append(headers)

# Зчитуємо дані з CSV-файлу та категоризуємо працівників
try:
    with open('employees2.csv', mode='r', encoding='utf-8') as csv_file:
        csv_reader = csv.DictReader(csv_file, delimiter=';')

        # Створюємо аркуші для різних вікових груп
        age_groups = {
            "younger_18": wb.create_sheet("younger_18"),  # Молодші 18 років
            "18-45": wb.create_sheet("18-45"),  # Вік 18-45 років
            "45-70": wb.create_sheet("45-70"),  # Вік 45-70 років
            "older_70": wb.create_sheet("older_70")  # Старші 70 років і більше
        }

        # Додаємо заголовки до аркушів вікових груп
        for sheet in age_groups.values():
            sheet.append(headers)

        # Отримуємо поточну дату для розрахунку віку
        today_date = date.today()

        # Перебираємо дані з CSV-файлу
        for row in csv_reader:
            # Розбираємо дату народження у форматі '%Y-%m-%d'
            birthdate = datetime.strptime(row['Date of Birth'], '%Y-%m-%d').date()
            age = calculate_age(birthdate)

            # Додаємо дані до аркушу "all"
            ws_all.append([row['Last Name'], row['First Name'], row['Middle Name'], row['Date of Birth'], age])

            # Категоризуємо дані за віковими групами
            if age < 18:
                age_groups["younger_18"].append(
                    [row['Last Name'], row['First Name'], row['Middle Name'], row['Date of Birth'], age])
            elif 18 <= age <= 45:
                age_groups["18-45"].append(
                    [row['Last Name'], row['First Name'], row['Middle Name'], row['Date of Birth'], age])
            elif 45 < age <= 70:
                age_groups["45-70"].append(
                    [row['Last Name'], row['First Name'], row['Middle Name'], row['Date of Birth'], age])
            else:
                age_groups["older_70"].append(
                    [row['Last Name'], row['First Name'], row['Middle Name'], row['Date of Birth'], age])
except FileNotFoundError:
    print("CSV-файл не знайдено.")
    exit(1)
except Exception as e:
    print(f"Помилка читання CSV-файлу: {e}")
    exit(1)

# Зберігаємо файл XLSX
try:
    wb.save('employee_data.xlsx')
    print("Готово")
except Exception as e:
    print(f"Помилка збереження файлу XLSX: {e}")